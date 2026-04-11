"""K-APT 아파트 정보 수집 — CSV 기본정보 + V4 API 상세정보.

Phase 1: CSV(apt_detail_info_v4.csv) → apt_kapt_info 기본정보 적재
Phase 2: V4 상세정보 API → 주차/CCTV/충전기 등 보충

사용법:
  python -m batch.kapt.collect_kapt_info --phase 1    # CSV 적재
  python -m batch.kapt.collect_kapt_info --phase 2    # V4 상세 API
  python -m batch.kapt.collect_kapt_info --phase all   # 1+2 순차
"""

import argparse
import csv
import re
import time
from pathlib import Path

import requests

from batch.config import DATA_GO_KR_API_KEY, DATA_GO_KR_RATE
from batch.db import get_connection, get_dict_cursor, query_all
from batch.logger import setup_logger

V4_DETAIL_URL = "https://apis.data.go.kr/1613000/AptBasisInfoServiceV4/getAphusDtlInfoV4"


def _norm_addr(addr: str) -> str:
    if not addr:
        return ""
    return re.sub(r"\(.*?\)", "", re.sub(r"\s+", " ", addr)).strip()


def _build_addr_index(conn) -> dict[str, str]:
    """전체 아파트의 도로명주소 → pnu 인덱스."""
    rows = query_all(conn,
        "SELECT pnu, new_plat_plc FROM apartments WHERE group_pnu = pnu AND new_plat_plc IS NOT NULL AND LENGTH(new_plat_plc) > 5")
    idx = {}
    for r in rows:
        na = _norm_addr(r["new_plat_plc"])
        if na:
            idx[na] = r["pnu"]
    return idx


# ── Phase 1: CSV 적재 ──

def phase1_csv(conn, logger):
    """CSV 기본정보 → apt_kapt_info 적재 + apartments 세대수 보정."""
    csv_path = Path(__file__).resolve().parents[2] / "apt_eda" / "data" / "raw" / "apt_detail_info_v4.csv"
    if not csv_path.exists():
        logger.error(f"CSV 없음: {csv_path}")
        return 0

    # kapt_code → pnu 인덱스
    kapt_code_idx = {}
    for r in query_all(conn,
            "SELECT pnu, kapt_code FROM apt_kapt_info WHERE kapt_code IS NOT NULL AND kapt_code != ''"):
        kapt_code_idx[r["kapt_code"]] = r["pnu"]

    cur = get_dict_cursor(conn)

    loaded = 0
    new_cnt = 0
    hhld_fixed = 0

    with open(csv_path, encoding="utf-8-sig") as f:
        for row in csv.DictReader(f):
            kapt_code = row.get("kaptCode", "").strip()
            if not kapt_code:
                continue

            pnu = kapt_code_idx.get(kapt_code)
            is_new = pnu is None
            if is_new:
                pnu = f"KAPT_{kapt_code}"

            kapt_name = row.get("kaptName", "").strip()
            sgg_code = pnu[:5] if not is_new else None

            def _int(val):
                try:
                    return int(float(val)) if val else None
                except (ValueError, TypeError):
                    return None

            def _float(val):
                try:
                    return float(val) if val else None
                except (ValueError, TypeError):
                    return None

            hhld = _int(row.get("hoCnt")) or 0
            dong = _int(row.get("kaptDongCnt")) or 0
            top_flr = _int(row.get("ktownFlrNo")) or 0
            total_area = _float(row.get("kaptTarea")) or 0
            priv_area = _float(row.get("privArea")) or 0
            elevator = _int(row.get("kaptdEcntp")) or 0
            use_date = str(row.get("kaptUsedate", "")).split(".")[0]

            # apt_kapt_info UPSERT (CSV 전체 컬럼)
            cur.execute("""
                INSERT INTO apt_kapt_info (
                    pnu, kapt_code, kapt_name, sigungu_code,
                    sale_type, heat_type, builder, developer,
                    apt_type, mgr_type, hall_type,
                    total_area, priv_area, mgmt_area, elevator_cnt,
                    ho_cnt, dong_cnt, top_floor, base_floor, use_date,
                    jibun_addr, road_addr, tel, fax, homepage, zipcode,
                    area_under_60, area_60_85, area_85_135, area_over_135
                ) VALUES (
                    %s, %s, %s, %s,
                    %s, %s, %s, %s,
                    %s, %s, %s,
                    %s, %s, %s, %s,
                    %s, %s, %s, %s, %s,
                    %s, %s, %s, %s, %s, %s,
                    %s, %s, %s, %s
                )
                ON CONFLICT (pnu) DO UPDATE SET
                    kapt_code = EXCLUDED.kapt_code,
                    kapt_name = COALESCE(EXCLUDED.kapt_name, apt_kapt_info.kapt_name),
                    sigungu_code = COALESCE(EXCLUDED.sigungu_code, apt_kapt_info.sigungu_code),
                    sale_type = EXCLUDED.sale_type, heat_type = EXCLUDED.heat_type,
                    builder = EXCLUDED.builder, developer = EXCLUDED.developer,
                    apt_type = EXCLUDED.apt_type, mgr_type = EXCLUDED.mgr_type,
                    hall_type = EXCLUDED.hall_type,
                    total_area = EXCLUDED.total_area, priv_area = EXCLUDED.priv_area,
                    mgmt_area = EXCLUDED.mgmt_area, elevator_cnt = EXCLUDED.elevator_cnt,
                    ho_cnt = EXCLUDED.ho_cnt, dong_cnt = EXCLUDED.dong_cnt,
                    top_floor = EXCLUDED.top_floor, base_floor = EXCLUDED.base_floor,
                    use_date = EXCLUDED.use_date,
                    jibun_addr = EXCLUDED.jibun_addr, road_addr = EXCLUDED.road_addr,
                    tel = EXCLUDED.tel, fax = EXCLUDED.fax,
                    homepage = EXCLUDED.homepage, zipcode = EXCLUDED.zipcode,
                    area_under_60 = EXCLUDED.area_under_60, area_60_85 = EXCLUDED.area_60_85,
                    area_85_135 = EXCLUDED.area_85_135, area_over_135 = EXCLUDED.area_over_135,
                    updated_at = NOW()
            """, [
                pnu, kapt_code, kapt_name or None, sgg_code,
                row.get("codeSaleNm", ""), row.get("codeHeatNm", ""),
                row.get("kaptBcompany", ""), row.get("kaptAcompany", ""),
                row.get("codeAptNm", ""), row.get("codeMgrNm", ""),
                row.get("codeHallNm", ""),
                total_area if total_area > 0 else None,
                priv_area if priv_area > 0 else None,
                _float(row.get("kaptMarea")),
                elevator if elevator > 0 else None,
                hhld if hhld > 0 else None,
                dong if dong > 0 else None,
                top_flr if top_flr > 0 else None,
                _int(row.get("kaptBaseFloor")),
                use_date if use_date and use_date != "0" else None,
                row.get("kaptAddr", "").strip() or None,
                row.get("doroJuso", "").strip() or None,
                row.get("kaptTel", "").strip() or None,
                row.get("kaptFax", "").strip() or None,
                row.get("kaptUrl", "").strip() or None,
                row.get("zipcode", "").strip() or None,
                _int(row.get("kaptMparea60")),
                _int(row.get("kaptMparea85")),
                _int(row.get("kaptMparea135")),
                _int(row.get("kaptMparea136")),
            ])
            loaded += 1
            if is_new:
                new_cnt += 1

            # apartments 세대수 보정 (기존 아파트만)
            if not is_new and hhld > 0:
                cur.execute("""
                    UPDATE apartments SET
                        total_hhld_cnt = GREATEST(COALESCE(total_hhld_cnt, 0), %s),
                        dong_count = GREATEST(COALESCE(dong_count, 0), %s),
                        max_floor = GREATEST(COALESCE(max_floor, 0), %s),
                        use_apr_day = COALESCE(NULLIF(use_apr_day, ''), %s)
                    WHERE pnu = %s AND COALESCE(total_hhld_cnt, 0) < %s
                """, [hhld, dong, top_flr, use_date, pnu, hhld])
                if cur.rowcount > 0:
                    hhld_fixed += 1

    conn.commit()
    logger.info(f"Phase 1 완료: {loaded}건 (기존갱신={loaded - new_cnt}, 신규={new_cnt}, 세대수보정={hhld_fixed})")
    return loaded


# ── Phase 1-B: 엑셀 기본정보 적재 ──

def phase1_xlsx(conn, logger):
    """K-APT 단지 기본정보 엑셀 → apt_kapt_info 적재 (CSV보다 상세)."""
    try:
        import openpyxl
    except ImportError:
        logger.error("openpyxl 미설치")
        return 0

    xlsx_path = Path(__file__).resolve().parents[2] / "apt_eda" / "data" / "k-apt" / "20260410_단지_기본정보.xlsx"
    if not xlsx_path.exists():
        logger.error(f"엑셀 없음: {xlsx_path}")
        return 0

    # kapt_code → pnu 인덱스
    kapt_code_idx = {}
    for r in query_all(conn,
            "SELECT pnu, kapt_code FROM apt_kapt_info WHERE kapt_code IS NOT NULL AND kapt_code != ''"):
        kapt_code_idx[r["kapt_code"]] = r["pnu"]

    cur = get_dict_cursor(conn)

    wb = openpyxl.load_workbook(xlsx_path, read_only=False)
    ws = wb.active
    headers = [cell.value for cell in ws[2]]  # 2행이 실제 헤더 (1행은 공지 병합셀)

    def _int(val):
        try:
            return int(float(val)) if val else None
        except (ValueError, TypeError):
            return None

    def _str(val):
        return str(val).strip() if val else None

    loaded = 0
    new_cnt = 0
    hhld_fixed = 0

    for row_cells in ws.iter_rows(min_row=3, values_only=True):
        row = dict(zip(headers, row_cells))

        kapt_code = _str(row.get("단지코드"))
        if not kapt_code:
            continue

        pnu = kapt_code_idx.get(kapt_code)
        is_new = pnu is None
        if is_new:
            pnu = f"KAPT_{kapt_code}"
        kapt_name = _str(row.get("단지명"))
        sgg_code = pnu[:5] if not is_new else None

        hhld = _int(row.get("세대수")) or 0
        dong = _int(row.get("동수")) or 0
        top_flr = _int(row.get("최고층수")) or 0
        use_date = _str(row.get("사용승인일"))

        cur.execute("""
            INSERT INTO apt_kapt_info (
                pnu, kapt_code, kapt_name, sigungu_code,
                sale_type, heat_type, builder, developer, apt_type, mgr_type, hall_type, structure,
                total_area, priv_area, mgmt_area, elevator_cnt,
                ho_cnt, dong_cnt, top_floor, top_floor_official, base_floor, use_date,
                sale_ho_cnt, rent_ho_cnt, rent_public_cnt, rent_private_cnt,
                area_under_60, area_60_85, area_85_135, area_over_135,
                mgmt_company, general_mgmt_type, general_mgmt_staff,
                security_type, security_staff, security_company,
                parking_cnt, parking_ground, parking_underground,
                total_car_cnt, ev_car_cnt,
                ev_charger_cnt, ev_charger_ground, ev_charger_underground,
                ev_parking_ground, ev_parking_underground,
                cctv_cnt, elevator_passenger, elevator_freight, elevator_mixed,
                elevator_disabled, elevator_emergency,
                home_network, welfare, convenience_facilities,
                jibun_addr, road_addr, tel, fax, zipcode
            ) VALUES (
                %s,%s,%s,%s,
                %s,%s,%s,%s,%s,%s,%s,%s,
                %s,%s,%s,%s,
                %s,%s,%s,%s,%s,%s,
                %s,%s,%s,%s,
                %s,%s,%s,%s,
                %s,%s,%s,
                %s,%s,%s,
                %s,%s,%s,
                %s,%s,
                %s,%s,%s,
                %s,%s,
                %s,%s,%s,%s,
                %s,%s,
                %s,%s,%s,
                %s,%s,%s,%s,%s
            )
            ON CONFLICT (pnu) DO UPDATE SET
                kapt_code = EXCLUDED.kapt_code,
                kapt_name = COALESCE(EXCLUDED.kapt_name, apt_kapt_info.kapt_name),
                sigungu_code = COALESCE(EXCLUDED.sigungu_code, apt_kapt_info.sigungu_code),
                sale_type = EXCLUDED.sale_type, heat_type = EXCLUDED.heat_type,
                builder = EXCLUDED.builder, developer = EXCLUDED.developer,
                apt_type = EXCLUDED.apt_type, mgr_type = EXCLUDED.mgr_type,
                hall_type = EXCLUDED.hall_type, structure = EXCLUDED.structure,
                ho_cnt = EXCLUDED.ho_cnt, dong_cnt = EXCLUDED.dong_cnt,
                top_floor = EXCLUDED.top_floor, top_floor_official = EXCLUDED.top_floor_official,
                base_floor = EXCLUDED.base_floor, use_date = EXCLUDED.use_date,
                sale_ho_cnt = EXCLUDED.sale_ho_cnt, rent_ho_cnt = EXCLUDED.rent_ho_cnt,
                rent_public_cnt = EXCLUDED.rent_public_cnt, rent_private_cnt = EXCLUDED.rent_private_cnt,
                mgmt_company = EXCLUDED.mgmt_company,
                general_mgmt_type = EXCLUDED.general_mgmt_type, general_mgmt_staff = EXCLUDED.general_mgmt_staff,
                security_type = EXCLUDED.security_type, security_staff = EXCLUDED.security_staff,
                security_company = EXCLUDED.security_company,
                parking_cnt = EXCLUDED.parking_cnt, parking_ground = EXCLUDED.parking_ground,
                parking_underground = EXCLUDED.parking_underground,
                total_car_cnt = EXCLUDED.total_car_cnt, ev_car_cnt = EXCLUDED.ev_car_cnt,
                ev_charger_cnt = EXCLUDED.ev_charger_cnt,
                ev_charger_ground = EXCLUDED.ev_charger_ground, ev_charger_underground = EXCLUDED.ev_charger_underground,
                ev_parking_ground = EXCLUDED.ev_parking_ground, ev_parking_underground = EXCLUDED.ev_parking_underground,
                cctv_cnt = EXCLUDED.cctv_cnt,
                elevator_passenger = EXCLUDED.elevator_passenger, elevator_freight = EXCLUDED.elevator_freight,
                elevator_mixed = EXCLUDED.elevator_mixed, elevator_disabled = EXCLUDED.elevator_disabled,
                elevator_emergency = EXCLUDED.elevator_emergency,
                home_network = EXCLUDED.home_network, welfare = EXCLUDED.welfare,
                convenience_facilities = EXCLUDED.convenience_facilities,
                jibun_addr = EXCLUDED.jibun_addr, road_addr = EXCLUDED.road_addr,
                tel = EXCLUDED.tel, fax = EXCLUDED.fax, zipcode = EXCLUDED.zipcode,
                updated_at = NOW()
        """, [
            pnu, kapt_code, kapt_name, sgg_code,
            _str(row.get("분양형태")), _str(row.get("난방방식")),
            _str(row.get("시공사")), _str(row.get("시행사")),
            _str(row.get("단지분류")), _str(row.get("관리방식")),
            _str(row.get("복도유형")), _str(row.get("건물구조")),
            None, None, None,  # total_area, priv_area, mgmt_area (엑셀에 없음, CSV/API에서)
            sum(filter(None, [
                _int(row.get("승강기(승객용)")), _int(row.get("승강기(화물용)")),
                _int(row.get("승강기(승객+화물)")), _int(row.get("승강기(장애인)")),
                _int(row.get("승강기(비상용)")), _int(row.get("승강기(기타)")),
            ])) or None,
            hhld if hhld > 0 else None,
            dong if dong > 0 else None,
            top_flr if top_flr > 0 else None,
            _int(row.get("최고층수(건축물대장상)")),
            _int(row.get("지하층수")),
            use_date,
            _int(row.get("분양세대수")), _int(row.get("임대세대수")),
            _int(row.get("임대세대수(공공)")), _int(row.get("임대세대수(민간)")),
            None, None, None, None,  # area_under_60 ~ area_over_135 (CSV에서)
            _str(row.get("주택관리업자")),
            _str(row.get("일반관리-관리방식")), _int(row.get("일반관리-인원")),
            _str(row.get("경비관리-관리방식")), _int(row.get("경비관리-인원")),
            _str(row.get("경비관리-계약업체")),
            _int(row.get("총주차대수")), _int(row.get("지상주차대수")), _int(row.get("지하주차대수")),
            _int(row.get("차량보유대수(전체)")), _int(row.get("차량보유대수(전기차)")),
            (_int(row.get("전기차 충전시설 설치대수(지상)")) or 0) + (_int(row.get("전기차 충전시설 설치대수(지하)")) or 0) or None,
            _int(row.get("전기차 충전시설 설치대수(지상)")), _int(row.get("전기차 충전시설 설치대수(지하)")),
            _int(row.get("전기차전용주차면수(지상)")), _int(row.get("전기차전용주차면수(지하)")),
            _int(row.get("CCTV대수")),
            _int(row.get("승강기(승객용)")), _int(row.get("승강기(화물용)")),
            _int(row.get("승강기(승객+화물)")), _int(row.get("승강기(장애인)")),
            _int(row.get("승강기(비상용)")),
            _str(row.get("홈네트워크")),
            _str(row.get("부대복리시설")), _str(row.get("입주편의시설")),
            _str(row.get("법정동주소")), _str(row.get("도로명주소")),
            _str(row.get("관리사무소 연락처")), _str(row.get("관리사무소 팩스")),
            _str(row.get("우편번호")),
        ])
        loaded += 1
        if is_new:
            new_cnt += 1

        if not is_new and hhld > 0:
            cur.execute("""
                UPDATE apartments SET
                    total_hhld_cnt = GREATEST(COALESCE(total_hhld_cnt, 0), %s),
                    dong_count = GREATEST(COALESCE(dong_count, 0), %s),
                    max_floor = GREATEST(COALESCE(max_floor, 0), %s),
                    use_apr_day = COALESCE(NULLIF(use_apr_day, ''), %s)
                WHERE pnu = %s AND COALESCE(total_hhld_cnt, 0) < %s
            """, [hhld, dong, top_flr, use_date, pnu, hhld])
            if cur.rowcount > 0:
                hhld_fixed += 1

        if loaded % 2000 == 0:
            conn.commit()
            logger.info(f"  진행: {loaded}건")

    wb.close()
    conn.commit()
    logger.info(f"Phase 1-B 완료: {loaded}건 (기존갱신={loaded - new_cnt}, 신규={new_cnt}, 세대수보정={hhld_fixed})")
    return loaded


# ── V4 상세정보 파싱 ──

def _parse_detail_item(item: dict) -> dict:
    """V4 상세정보 API 응답 → DB 저장용 dict 변환."""
    parking = int(item.get("kaptdPcntu") or 0)
    cctv = int(item.get("kaptdCccnt") or 0)
    ev_ground = int(item.get("groundElChargerCnt") or 0)
    ev_under = int(item.get("undergroundElChargerCnt") or 0)
    structure = item.get("codeStr", "")
    subway_line = item.get("subwayLine") or ""
    subway_station = item.get("subwayStation") or ""
    subway_info = f"{subway_line} {subway_station}".strip() if subway_line or subway_station else None
    bus_time = item.get("kaptdWtimebus", "")
    welfare = item.get("welfareFacility", "")
    return {
        "parking_cnt": parking or None,
        "cctv_cnt": cctv or None,
        "ev_charger_cnt": (ev_ground + ev_under) or None,
        "structure": structure or None,
        "subway_info": subway_info,
        "bus_time": bus_time or None,
        "welfare": welfare or None,
    }


def _fetch_detail(kapt_code: str) -> dict | None:
    """V4 상세정보 API 호출 → item dict 반환. 실패 시 None."""
    try:
        resp = requests.get(V4_DETAIL_URL, params={
            "serviceKey": DATA_GO_KR_API_KEY,
            "kaptCode": kapt_code,
            "type": "json",
        }, timeout=10)
        resp.raise_for_status()
        return resp.json().get("response", {}).get("body", {}).get("item", {}) or None
    except Exception:
        return None


# ── Phase 2: V4 상세정보 API (미수집 건) ──

def phase2_api(conn, logger):
    """kapt_code가 있는 아파트 → V4 상세정보 API로 주차/CCTV 등 보충."""
    if not DATA_GO_KR_API_KEY:
        logger.error("DATA_GO_KR_API_KEY 미설정")
        return 0

    cur = get_dict_cursor(conn)
    targets = query_all(conn,
        "SELECT pnu, kapt_code FROM apt_kapt_info WHERE kapt_code IS NOT NULL AND kapt_code != '' AND parking_cnt IS NULL")

    logger.info(f"Phase 2 대상: {len(targets)}건")
    updated = 0
    failed = 0

    for i, row in enumerate(targets):
        item = _fetch_detail(row["kapt_code"])
        time.sleep(DATA_GO_KR_RATE)

        if not item:
            failed += 1
            continue

        vals = _parse_detail_item(item)
        cur.execute("""
            UPDATE apt_kapt_info SET
                parking_cnt = %s, cctv_cnt = %s, ev_charger_cnt = %s,
                structure = %s, subway_info = %s, bus_time = %s, welfare = %s,
                updated_at = NOW()
            WHERE pnu = %s
        """, [vals["parking_cnt"], vals["cctv_cnt"], vals["ev_charger_cnt"],
              vals["structure"], vals["subway_info"], vals["bus_time"], vals["welfare"],
              row["pnu"]])
        updated += 1

        if (i + 1) % 500 == 0:
            conn.commit()
            logger.info(f"  진행: {i+1}/{len(targets)} (보충={updated}, 실패={failed})")

    conn.commit()
    logger.info(f"Phase 2 완료: 보충={updated}, 실패={failed}")
    return updated


# ── 기본정보 파싱 ──

def _parse_basic_item(item: dict) -> dict:
    """V4 기본정보 API 응답 → DB 저장용 dict 변환."""
    def _int(val):
        try:
            return int(float(val)) if val else None
        except (ValueError, TypeError):
            return None

    def _float(val):
        try:
            return float(val) if val else None
        except (ValueError, TypeError):
            return None

    use_date = str(item.get("kaptUsedate", "")).split(".")[0]
    return {
        "kapt_name": (item.get("kaptName") or "").strip() or None,
        "sale_type": item.get("codeSaleNm", ""),
        "heat_type": item.get("codeHeatNm", ""),
        "builder": item.get("kaptBcompany", ""),
        "developer": item.get("kaptAcompany", ""),
        "apt_type": item.get("codeAptNm", ""),
        "mgr_type": item.get("codeMgrNm", ""),
        "hall_type": item.get("codeHallNm", ""),
        "total_area": _float(item.get("kaptTarea")),
        "priv_area": _float(item.get("privArea")),
        "mgmt_area": _float(item.get("kaptMarea")),
        "elevator_cnt": _int(item.get("kaptdEcntp")),
        "ho_cnt": _int(item.get("hoCnt")),
        "dong_cnt": _int(item.get("kaptDongCnt")),
        "top_floor": _int(item.get("ktownFlrNo")),
        "base_floor": _int(item.get("kaptBaseFloor")),
        "use_date": use_date if use_date and use_date != "0" else None,
        "jibun_addr": (item.get("kaptAddr") or "").strip() or None,
        "road_addr": (item.get("doroJuso") or "").strip() or None,
        "tel": (item.get("kaptTel") or "").strip() or None,
        "fax": (item.get("kaptFax") or "").strip() or None,
        "homepage": (item.get("kaptUrl") or "").strip() or None,
        "zipcode": (item.get("zipcode") or "").strip() or None,
        "area_under_60": _int(item.get("kaptMparea60")),
        "area_60_85": _int(item.get("kaptMparea85")),
        "area_85_135": _int(item.get("kaptMparea135")),
        "area_over_135": _int(item.get("kaptMparea136")),
    }


# ── Phase refresh: 전체 갱신 (기본 + 상세) ──

_BASIC_COLS = (
    "kapt_name", "sale_type", "heat_type", "builder", "developer",
    "apt_type", "mgr_type", "hall_type", "total_area", "priv_area", "mgmt_area",
    "elevator_cnt", "ho_cnt", "dong_cnt", "top_floor", "base_floor", "use_date",
    "jibun_addr", "road_addr", "tel", "fax", "homepage", "zipcode",
    "area_under_60", "area_60_85", "area_85_135", "area_over_135",
)
_DETAIL_COLS = ("parking_cnt", "cctv_cnt", "ev_charger_cnt", "structure", "subway_info", "bus_time", "welfare")
_ALL_COLS = _BASIC_COLS + _DETAIL_COLS


def phase2_refresh(conn, logger):
    """기존 kapt_info 전체 갱신 — 기본정보 + 상세정보 API, 변경분만 UPDATE."""
    if not DATA_GO_KR_API_KEY:
        logger.error("DATA_GO_KR_API_KEY 미설정")
        return 0

    cur = get_dict_cursor(conn)
    select_cols = ", ".join(_ALL_COLS)
    targets = query_all(conn,
        f"SELECT pnu, kapt_code, {select_cols} "
        f"FROM apt_kapt_info WHERE kapt_code IS NOT NULL AND kapt_code != ''")

    logger.info(f"Refresh 대상: {len(targets)}건")
    changed = 0
    skipped = 0
    failed = 0

    for i, row in enumerate(targets):
        kapt_code = row["kapt_code"]

        # 기본정보 API
        basic_item = _fetch_kapt_basic(kapt_code)
        time.sleep(DATA_GO_KR_RATE)

        # 상세정보 API
        detail_item = _fetch_detail(kapt_code)
        time.sleep(DATA_GO_KR_RATE)

        if not basic_item and not detail_item:
            failed += 1
            continue

        basic_vals = _parse_basic_item(basic_item) if basic_item else {}
        detail_vals = _parse_detail_item(detail_item) if detail_item else {}
        vals = {**basic_vals, **detail_vals}

        # 변경 감지: 기존값과 비교
        has_diff = any(row.get(col) != vals.get(col) for col in _ALL_COLS if col in vals)
        if not has_diff:
            skipped += 1
        else:
            set_parts = []
            params = []
            for col in _ALL_COLS:
                if col in vals:
                    set_parts.append(f"{col} = %s")
                    params.append(vals[col])
            set_parts.append("updated_at = NOW()")
            params.append(row["pnu"])

            cur.execute(
                f"UPDATE apt_kapt_info SET {', '.join(set_parts)} WHERE pnu = %s",
                params,
            )
            changed += 1

        if (i + 1) % 500 == 0:
            conn.commit()
            logger.info(f"  진행: {i+1}/{len(targets)} (갱신={changed}, 스킵={skipped}, 실패={failed})")

    conn.commit()
    logger.info(f"Refresh 완료: 갱신={changed}, 스킵={skipped}, 실패={failed}")
    return changed


# ── 신규 아파트 K-APT 매칭 (trade 배치 연동) ──

V4_BASIC_URL = "https://apis.data.go.kr/1613000/AptBasisInfoServiceV4/getAphusBassInfoV4"
APT_LIST_URL = "https://apis.data.go.kr/1613000/AptListService3/getTotalAptList3"


def _load_kapt_list() -> list[dict]:
    """K-APT 전체 목록 조회 (캐시 없이 매번 API 호출)."""
    all_items = []
    page = 1
    while True:
        try:
            resp = requests.get(APT_LIST_URL, params={
                "serviceKey": DATA_GO_KR_API_KEY, "numOfRows": "1000",
                "pageNo": str(page), "type": "json",
            }, timeout=30)
            data = resp.json().get("response", {}).get("body", {})
            items = data.get("items", [])
            if not items:
                break
            if isinstance(items, dict):
                items = [items]
            all_items.extend(items)
            if len(all_items) >= (data.get("totalCount") or 0):
                break
            page += 1
            time.sleep(DATA_GO_KR_RATE)
        except Exception:
            break
    return all_items


def _fetch_kapt_basic(kapt_code: str) -> dict:
    """V4 기본정보 API 호출."""
    try:
        resp = requests.get(V4_BASIC_URL, params={
            "serviceKey": DATA_GO_KR_API_KEY, "kaptCode": kapt_code, "type": "json",
        }, timeout=10)
        return resp.json().get("response", {}).get("body", {}).get("item", {})
    except Exception:
        return {}


def _fetch_kapt_detail(kapt_code: str) -> dict:
    """V4 상세정보 API 호출."""
    try:
        resp = requests.get(V4_DETAIL_URL, params={
            "serviceKey": DATA_GO_KR_API_KEY, "kaptCode": kapt_code, "type": "json",
        }, timeout=10)
        return resp.json().get("response", {}).get("body", {}).get("item", {})
    except Exception:
        return {}


def enrich_kapt_for_new(conn, logger, pnu_list: list[str]):
    """신규 등록 아파트에 대해 K-APT 정보 매칭 + 적재."""
    if not DATA_GO_KR_API_KEY or not pnu_list:
        return 0

    # 신규 아파트 주소 조회
    ph = ",".join(["%s"] * len(pnu_list))
    new_apts = query_all(conn,
        f"SELECT pnu, new_plat_plc FROM apartments WHERE pnu IN ({ph}) AND new_plat_plc IS NOT NULL",
        pnu_list)
    if not new_apts:
        return 0

    new_addr = {_norm_addr(r["new_plat_plc"]): r["pnu"] for r in new_apts if r["new_plat_plc"]}

    # K-APT 전체 목록에서 도로명주소 매칭용 인덱스 구축
    logger.info(f"  K-APT 목록 조회 중...")
    kapt_list = _load_kapt_list()
    logger.info(f"  K-APT 목록: {len(kapt_list)}건")

    # kaptCode별 doroJuso 매칭을 위해 기본정보 API로 주소 확인
    cur = get_dict_cursor(conn)
    matched = 0

    for kapt in kapt_list:
        kapt_code = kapt.get("kaptCode")
        if not kapt_code:
            continue

        # 기본정보 조회
        basic = _fetch_kapt_basic(kapt_code)
        time.sleep(DATA_GO_KR_RATE)
        if not basic:
            continue

        doro = _norm_addr(basic.get("doroJuso", ""))
        pnu = new_addr.get(doro)
        if not pnu:
            continue

        # 이미 apt_kapt_info에 있으면 스킵
        existing = query_all(conn, "SELECT pnu FROM apt_kapt_info WHERE pnu = %s", [pnu])
        if existing:
            continue

        # 상세정보 조회
        detail = _fetch_kapt_detail(kapt_code)
        time.sleep(DATA_GO_KR_RATE)

        # 기본정보 적재
        try:
            hhld = int(basic.get("hoCnt") or 0)
            dong = int(basic.get("kaptDongCnt") or 0)
            top_flr = int(basic.get("ktownFlrNo") or 0)
            total_area = float(basic.get("kaptTarea") or 0)
            priv_area = float(basic.get("privArea") or 0)
            elevator = int(basic.get("kaptdEcntp") or 0)
        except (ValueError, TypeError):
            hhld = dong = top_flr = elevator = 0
            total_area = priv_area = 0

        parking = int(detail.get("kaptdPcntu") or 0) if detail else 0
        cctv = int(detail.get("kaptdCccnt") or 0) if detail else 0
        ev_g = int(detail.get("groundElChargerCnt") or 0) if detail else 0
        ev_u = int(detail.get("undergroundElChargerCnt") or 0) if detail else 0
        structure = detail.get("codeStr", "") if detail else ""
        subway = f"{detail.get('subwayLine','')} {detail.get('subwayStation','')}".strip() if detail else None
        bus_time = detail.get("kaptdWtimebus", "") if detail else ""
        welfare = detail.get("welfareFacility", "") if detail else ""

        cur.execute("""
            INSERT INTO apt_kapt_info (pnu, kapt_code, sale_type, heat_type, builder, developer,
                apt_type, mgr_type, hall_type, structure, total_area, priv_area,
                parking_cnt, cctv_cnt, elevator_cnt, ev_charger_cnt, subway_info, bus_time, welfare)
            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
            ON CONFLICT (pnu) DO UPDATE SET
                kapt_code=EXCLUDED.kapt_code, sale_type=EXCLUDED.sale_type,
                heat_type=EXCLUDED.heat_type, builder=EXCLUDED.builder, developer=EXCLUDED.developer,
                parking_cnt=EXCLUDED.parking_cnt, cctv_cnt=EXCLUDED.cctv_cnt,
                elevator_cnt=EXCLUDED.elevator_cnt, ev_charger_cnt=EXCLUDED.ev_charger_cnt,
                structure=EXCLUDED.structure, updated_at=NOW()
        """, [pnu, kapt_code,
              basic.get("codeSaleNm",""), basic.get("codeHeatNm",""),
              basic.get("kaptBcompany",""), basic.get("kaptAcompany",""),
              basic.get("codeAptNm",""), basic.get("codeMgrNm",""),
              basic.get("codeHallNm",""), structure or None,
              total_area or None, priv_area or None,
              parking or None, cctv or None, elevator or None,
              (ev_g + ev_u) or None, subway, bus_time or None, welfare or None])

        # apartments 세대수 보정
        if hhld > 0:
            cur.execute("""
                UPDATE apartments SET
                    total_hhld_cnt = GREATEST(COALESCE(total_hhld_cnt,0), %s),
                    dong_count = GREATEST(COALESCE(dong_count,0), %s),
                    max_floor = GREATEST(COALESCE(max_floor,0), %s)
                WHERE pnu = %s
            """, [hhld, dong, top_flr, pnu])

        matched += 1
        if matched <= 3 or matched % 10 == 0:
            logger.info(f"  K-APT 매칭: {basic.get('kaptName','')} → {pnu}")

    conn.commit()
    logger.info(f"  K-APT 신규 매칭: {matched}건")
    return matched


def main():
    parser = argparse.ArgumentParser(description="K-APT 아파트 정보 수집")
    parser.add_argument("--phase", required=True, choices=["1", "xlsx", "2", "all", "refresh"],
                        help="1: CSV 적재, xlsx: 엑셀 적재 (상세), 2: V4 상세 API (미수집), all: 1+xlsx+2, refresh: 전체 갱신")
    args = parser.parse_args()

    logger = setup_logger("kapt_info")
    conn = get_connection()

    try:
        if args.phase in ("1", "all"):
            phase1_csv(conn, logger)
        if args.phase in ("xlsx", "all"):
            phase1_xlsx(conn, logger)
        if args.phase in ("2", "all"):
            phase2_api(conn, logger)
        if args.phase == "refresh":
            phase2_refresh(conn, logger)
    finally:
        conn.close()


if __name__ == "__main__":
    main()
